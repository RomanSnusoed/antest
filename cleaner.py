# Импорты
import fitz  # PyMuPDF
import pytesseract
from PIL import Image, ImageEnhance
import spacy
import re
import io
import os
import docx  # Для работы с .docx файлами
from striprtf.striprtf import rtf_to_text  # Для работы с .rtf файлами
import csv  # Для работы с .csv файлами
from openpyxl import load_workbook  # Для работы с .xlsx файлами
from bs4 import BeautifulSoup  # Для работы с .html файлами
import win32com.client  # Для работы с .doc файлами
from odf.opendocument import load  # Для работы с .odt файлами
from odf.text import P  # Для извлечения текста из .odt

# Укажите путь к Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Загрузка модели Spacy для обработки текста
nlp = spacy.load("en_core_web_trf")

# Функции
def preprocess_image(image_path):
    """
    Предварительная обработка изображения для улучшения качества OCR.
    """
    img = Image.open(image_path)
    img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)  # Увеличение размера
    img = img.convert('L')  # Преобразование в градации серого

    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(4)  # Увеличение контраста
    enhancer = ImageEnhance.Brightness(img)
    img = enhancer.enhance(1.5)  # Увеличение яркости

    return img

def correct_image_orientation(image):
    """
    Определяет ориентацию изображения и поворачивает его в правильную ориентацию.
    """
    try:
        osd = pytesseract.image_to_osd(image)
        rotation_angle = int(re.search(r"Rotate: (\d+)", osd).group(1))
        if rotation_angle != 0:
            print(f"Поворачиваем изображение на {rotation_angle} градусов")
            image = image.rotate(-rotation_angle, expand=True)
    except Exception as e:
        print(f"Ошибка при определении ориентации: {e}")
    return image

def extract_text_from_image(image_path):
    """
    Извлекает текст из изображения (JPG или PNG) с помощью Tesseract OCR.
    """
    try:
        img = preprocess_image(image_path)
        # Убираем вызов correct_image_orientation для JPG
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(img, config=custom_config)
        return text.strip()
    except Exception as e:
        print(f"Ошибка при извлечении текста из изображения: {e}")
        return ""

def extract_text_from_pdf(pdf_path):
    """
    Извлекает текст из PDF. Если текст не извлекается, извлекает изображения и применяет OCR.
    """
    text = ""
    try:
        # Открываем PDF
        pdf_document = fitz.open(pdf_path)
        for page_number in range(len(pdf_document)):
            page = pdf_document[page_number]
            
            # Попытка извлечь текстовый слой
            page_text = page.get_text()
            if page_text.strip():
                text += page_text + "\n"
        
        # Если текст извлечён, возвращаем его
        if text.strip():
            return text
    except Exception as e:
        print(f"Ошибка при извлечении текста из PDF: {e}")

    # Если текст не извлечён, извлекаем изображения
    print("Текст не найден. Извлекаем изображения для OCR...")
    os.makedirs("processed_images", exist_ok=True)  # Убедимся, что папка существует
    for page_number in range(len(pdf_document)):
        page = pdf_document[page_number]
        images = page.get_images(full=True)
        for img_index, img in enumerate(images):
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]
            image = Image.open(io.BytesIO(image_bytes))
            
            # Корректируем ориентацию изображения
            image = correct_image_orientation(image)
            
            # Сохраняем изображение для отладки
            image_path = f"processed_images/page_{page_number + 1}_image_{img_index + 1}.jpg"
            image.save(image_path, "JPEG")
            print(f"Изображение сохранено: {image_path}")
            
            # Применяем OCR к изображению
            text += pytesseract.image_to_string(image) + "\n"
    return text

def extract_text_from_docx(docx_path):
    """
    Извлекает текст из .docx файла.
    """
    try:
        doc = docx.Document(docx_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Ошибка при извлечении текста из .docx файла: {e}")
        return ""

def extract_text_from_doc(doc_path):
    """
    Извлекает текст из .doc файла с помощью pywin32.
    """
    try:
        word = win32com.client.Dispatch("Word.Application")
        doc = word.Documents.Open(doc_path)
        text = doc.Content.Text
        doc.Close()
        word.Quit()
        return text
    except Exception as e:
        print(f"Ошибка при извлечении текста из .doc файла: {e}")
        return ""

def extract_text_from_odt(odt_path):
    """
    Извлекает текст из .odt файла.
    """
    try:
        doc = load(odt_path)
        paragraphs = doc.getElementsByType(P)
        text = ""
        for paragraph in paragraphs:
            if paragraph.firstChild:  # Проверяем, есть ли дочерний элемент
                text += str(paragraph.firstChild) + "\n"
        return text
    except Exception as e:
        print(f"Ошибка при извлечении текста из .odt файла: {e}")
        return ""

def clean_and_remove_personal_info(text):
    """
    Removes personal information using Spacy and regex,
    but preserves data recognized by Spacy as QUANTITY, PERCENT, and CARDINAL.
    """
    doc = nlp(text)
    cleaned_text = text
    regex_matches = []

    # Extract all entities labeled as QUANTITY, PERCENT, and CARDINAL
    quantities = [ent.text for ent in doc.ents if ent.label_ == "QUANTITY"]
    percents = [ent.text for ent in doc.ents if ent.label_ == "PERCENT"]
    cardinals = [ent.text for ent in doc.ents if ent.label_ == "CARDINAL"]
    print(f"Found QUANTITY entities: {quantities}")
    print(f"Found PERCENT entities: {percents}")
    print(f"Found CARDINAL entities: {cardinals}")

    # Temporarily replace QUANTITY, PERCENT, and CARDINAL entities with unique placeholders
    for i, quantity in enumerate(quantities):
        placeholder = f"__QUANTITY_{i}__"
        cleaned_text = cleaned_text.replace(quantity, placeholder)

    for i, percent in enumerate(percents):
        placeholder = f"__PERCENT_{i}__"
        cleaned_text = cleaned_text.replace(percent, placeholder)

    for i, cardinal in enumerate(cardinals):
        placeholder = f"__CARDINAL_{i}__"
        # Проверяем, является ли CARDINAL чувствительным
        if len(cardinal) > 4 or re.match(r'\b\d{5}\b|\b\d{3}[-\s]?\d{3}[-\s]?\d{4}\b', cardinal):  # Длинные числа или номера телефонов
            print(f"Removing sensitive CARDINAL: {cardinal}")
            # Удаляем чувствительный CARDINAL
            cleaned_text = cleaned_text.replace(placeholder, "[REDACTED]")
        else:
            # Восстанавливаем не чувствительные CARDINAL
            print(f"Restoring non-sensitive CARDINAL: {cardinal}")
            cleaned_text = cleaned_text.replace(placeholder, cardinal)

    # Regex patterns for removing sensitive information
    patterns = [
        r'\b\d{6,}\b',  # Numbers with 6 or more digits
        r'\d{2}/\d{2}/\d{4}',  # Dates in DD/MM/YYYY format
        r'\b[A-Z]{1,2}\d{6,}\b',  # Identifiers (e.g., passport numbers)
        r'\b\d{3}[-\s]?\d{3}[-\s]?\d{4}\b',  # Phone numbers (e.g., 123-456-7890)
        r'\+?\d{1,3}[-\s]?\(?\d{1,4}\)?[-\s]?\d{1,4}[-\s]?\d{1,4}',  # International phone numbers
        r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',  # Email addresses
        r'\bwww\.[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b',  # Websites
        r'\bhttp[s]?://[^\s]+',  # URLs
        r'\b\d{1,2}-\d{1,2}-\d{4}\b',  # Dates in DD-MM-YYYY format
        r'\b\d{1,2}\.\d{1,2}\.\d{4}\b',  # Dates in DD.MM.YYYY format
        r'\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b',  # Dates in DD MMM YYYY format
        r'\b[A-Za-z]{3,9}\s+\d{4}\b',  # Dates in MMM YYYY format
        r'[A-Za-z]+\s+Court',  # Mentions of courts
        r'Managing Director\s+[A-Za-z\s]+',  # Mentions of directors
        r'\b[A-Za-zäöüÄÖÜß]+\s+\d{1,4},?\s+[A-Za-zäöüÄÖÜß]+\s+\d{5}\b',  # Addresses (street, number, city, postal code)
        r'\b\d{3,5}/\d{3,5}/\d{3,5}\b',  # Tax numbers (e.g., 123/456/789)
        r'Fax[:\s-]*\d{1,4}[-\s]?\d{1,4}[-\s]?\d{1,4}',  # Fax numbers
        r'[A-Za-zäöüÄÖÜß]+\s+\d{1,4},?\s+D-\d{5}',  # Addresses with postal codes
        r'[A-Za-zäöüÄÖÜß\-]+\s+\d{1,4}',  # Streets with house numbers
        r'\b\d{1,4}\s+[A-Za-z\s]+,\s*[A-Za-z\s]*\b',  # General address format (e.g., "20 HARRIER PARK")
        r'\bNHS number[^\n]*',  # Удаление строк с NHS number
        r'\bGMC[:\s]*\d{6,}\b',  # Удаление GMC номера (например, "GMC: 7774760")
        r'\bGMC[:\s]*\d{6,}.*?\)',  # Удаляет GMC номер вместе с текстом до закрывающей скобки
        r'\bGMC[:\s]*\d{6,}.*?(?=\)|$)',  # Удаляет GMC номер вместе с текстом до закрывающей скобки или конца строки
        r'\bGMC[:\s]*\d{6,}.*?(?=\)|$)',  # Удаляет GMC номер вместе с текстом до закрывающей скобки или конца строки
        r'\bDr\s+[A-Za-z]+\b.*?(?=\n|$)',  # Удаляет строки с упоминанием "Dr" и текстом после
    ]

    for pattern in patterns:
        matches = re.findall(pattern, cleaned_text)
        if matches:
            regex_matches.extend(matches)
            for match in matches:
                cleaned_text = cleaned_text.replace(match, "[REDACTED]")




def clean_and_remove_personal_info(text, return_regex_matches=False):
    """
    Removes personal information using regex patterns.
    If return_regex_matches is True, returns the matches instead of the cleaned text.
    """
    cleaned_text = text
    regex_matches = []

    # Regex patterns for removing sensitive information
    patterns = [
        r'\b\d{6,}\b',  # Numbers with 6 or more digits
        r'\d{2}/\d{2}/\d{4}',  # Dates in DD/MM/YYYY format
        r'\b[A-Z]{1,2}\d{6,}\b',  # Identifiers (e.g., passport numbers)
        r'\b\d{3}[-\s]?\d{3}[-\s]?\d{4}\b',  # Phone numbers (e.g., 123-456-7890)
        r'\+?\d{1,3}[-\s]?\(?\d{1,4}\)?[-\s]?\d{1,4}[-\s]?\d{1,4}',  # International phone numbers
        r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',  # Email addresses
    ]

    # Apply regex patterns to find and redact sensitive information
    for pattern in patterns:
        matches = re.findall(pattern, cleaned_text)
        if matches:
            regex_matches.extend(matches)
            for match in matches:
                cleaned_text = cleaned_text.replace(match, "[REDACTED]")

    # If return_regex_matches is True, return the matches and skip further processing
    if return_regex_matches:
        return regex_matches

    # Process Spacy entities
    doc = nlp(cleaned_text)
    quantities = [ent.text for ent in doc.ents if ent.label_ == "QUANTITY"]
    percents = [ent.text for ent in doc.ents if ent.label_ == "PERCENT"]
    cardinals = [ent.text for ent in doc.ents if ent.label_ == "CARDINAL"]

    print(f"Found QUANTITY entities: {quantities}")
    print(f"Found PERCENT entities: {percents}")
    print(f"Found CARDINAL entities: {cardinals}")

    # Temporarily replace QUANTITY, PERCENT, and CARDINAL entities with unique placeholders
    for i, quantity in enumerate(quantities):
        placeholder = f"__QUANTITY_{i}__"
        cleaned_text = cleaned_text.replace(quantity, placeholder)

    for i, percent in enumerate(percents):
        placeholder = f"__PERCENT_{i}__"
        cleaned_text = cleaned_text.replace(percent, placeholder)

    for i, cardinal in enumerate(cardinals):
        placeholder = f"__CARDINAL_{i}__"
        cleaned_text = cleaned_text.replace(cardinal, placeholder)

    # Restore QUANTITY, PERCENT, and CARDINAL back into the text
    for i, quantity in enumerate(quantities):
        placeholder = f"__QUANTITY_{i}__"
        cleaned_text = cleaned_text.replace(placeholder, quantity)

    for i, percent in enumerate(percents):
        placeholder = f"__PERCENT_{i}__"
        cleaned_text = cleaned_text.replace(placeholder, percent)

    for i, cardinal in enumerate(cardinals):
        placeholder = f"__CARDINAL_{i}__"
        # Check if CARDINAL is sensitive
        if re.match(r'^\d{5,}$', cardinal):  # Long numbers (e.g., identifiers)
            print(f"Removing sensitive CARDINAL: {cardinal}")
            cleaned_text = cleaned_text.replace(placeholder, "[REDACTED]")
        elif re.match(r'^\d{3}[-\s]?\d{3}[-\s]?\d{4}$', cardinal):  # Phone numbers
            print(f"Removing sensitive CARDINAL: {cardinal}")
            cleaned_text = cleaned_text.replace(placeholder, "[REDACTED]")
        elif re.match(r'^\d+(\.\d+)?-\d+(\.\d+)?$', cardinal):  # Ranges of numbers
            print(f"Restoring non-sensitive CARDINAL (range): {cardinal}")
            cleaned_text = cleaned_text.replace(placeholder, cardinal)
        elif re.match(r'^\d+(\.\d+)?$', cardinal):  # Simple numbers
            print(f"Restoring non-sensitive CARDINAL: {cardinal}")
            cleaned_text = cleaned_text.replace(placeholder, cardinal)
        else:
            print(f"Restoring non-sensitive CARDINAL (default): {cardinal}")
            cleaned_text = cleaned_text.replace(placeholder, cardinal)

    # Remove all entities found by Spacy except QUANTITY, PERCENT, and CARDINAL
    for ent in doc.ents:
        if ent.label_ not in ["QUANTITY", "PERCENT", "CARDINAL"]:
            print(f"Found by Spacy entity: {ent.text} ({ent.label_})")
            cleaned_text = cleaned_text.replace(ent.text, "[REDACTED]")

    return cleaned_text

if __name__ == "__main__":
    file_path = r"C:\Users\38066\Dr-Adem Dropbox\Team AI\Dummy Text files\001fake.pdf"

    
    if file_path.endswith(".pdf"):
        extracted_text = extract_text_from_pdf(file_path)
    elif file_path.endswith(".docx"):
        extracted_text = extract_text_from_docx(file_path)
    elif file_path.endswith(".doc"):
        extracted_text = extract_text_from_doc(file_path)
    elif file_path.endswith(".odt"):
        extracted_text = extract_text_from_odt(file_path)
    elif file_path.endswith((".jpg", ".jpeg", ".png", ".tiff")):
        extracted_text = extract_text_from_image(file_path)
    elif file_path.endswith(".txt"):
        with open(file_path, "r", encoding="utf-8") as file:
            extracted_text = file.read()
    elif file_path.endswith(".rtf"):
        with open(file_path, "r", encoding="utf-8") as file:
            rtf_content = file.read()
            extracted_text = rtf_to_text(rtf_content)
    elif file_path.endswith(".csv"):
        extracted_text = ""
        with open(file_path, "r", encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                extracted_text += " ".join(row) + "\n"
    elif file_path.endswith(".xlsx"):
        extracted_text = ""
        workbook = load_workbook(file_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                extracted_text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
    elif file_path.endswith(".html"):
        with open(file_path, "r", encoding="utf-8") as file:
            soup = BeautifulSoup(file, "html.parser")
            extracted_text = soup.get_text()
    else:
        print("Неподдерживаемый формат файла.")
        extracted_text = ""

    if extracted_text:
        print("Извлечённый текст:")
        print(extracted_text)
        cleaned_text = clean_and_remove_personal_info(extracted_text)
        print("\nОчищенный текст:")
        print(cleaned_text)